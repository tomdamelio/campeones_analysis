"""
Análisis Descriptivo — Señales Completas por Estímulo
======================================================
Objetivos:
  1. Describir registros de actividad electrodérmica (EDA) y sus features.
  2. Describir autorreportes continuos (momento a momento) de valencia y arousal.
     La señal de joystick [-1, 1] se transforma a escala SAM [1, 9] antes de
     cualquier agregación: y = (x + 1) * 4 + 1. Esto aplica a todas las métricas
     derivadas (media, mediana, SD, rango, derivada). No se aplica abs() en ningún
     caso; el punto neutro de la escala es 5.

Jerarquía de agregación:
  Nivel 1 — stimulus_level_data.csv     : una fila por sujeto × estímulo
  Nivel 2 — stimulus_agg_valence.xlsx   : agregado por estímulo (valencia)
           — stimulus_agg_arousal.xlsx  : agregado por estímulo (arousal)
           — stimulus_agg_eda.xlsx      : agregado EDA por estímulo
  Nivel 3 — quadrant_desc_valence.xlsx  : descriptivo por cuadrante (valencia)
           — quadrant_desc_arousal.xlsx : descriptivo por cuadrante (arousal)
           — quadrant_desc_eda.xlsx     : descriptivo EDA por cuadrante
  Overview— dataset_overview.xlsx       : métricas generales del dataset

ICC temporal: correlación intraclase (two-way mixed, consistency) sobre perfiles
  temporales remuestreados a grilla común. Se reporta ICC + corrección Spearman-Brown
  por duración y CV del estímulo como covariable interpretativa.

Clasificación teórica de estímulos (Russell, 1980):
  HVLA: [1, 5, 6]   HVHA: [2, 3, 4, 7]
  LVLA: [8, 11, 13] LVHA: [9, 10, 12, 14]

Outputs en: results/descriptive/
"""

# ─────────────────────────────────────────────────────────────────────────────
# IMPORTS
# ─────────────────────────────────────────────────────────────────────────────
import os, re, json, warnings
import numpy as np
import pandas as pd
import mne
import neurokit2 as nk
from scipy.signal import find_peaks, welch
from scipy.interpolate import interp1d
from scipy.stats import pearsonr
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

warnings.filterwarnings("ignore")

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN GLOBAL
# ─────────────────────────────────────────────────────────────────────────────
SUBJECT_IDS = [
    "19","20","21","22","23","24","25","26","27","28",
    "29","30","31","32","33","34","35","36","37","38",
    "39","40","42","43","46",
]

CHANNEL_TO_PLOT     = "joystick_x"
SMOOTHING_WINDOW_S  = 1.0       # ventana de suavizado para el autorreporte (seg)

# Transformación lineal joystick [-1, 1] → escala SAM [1, 9]
# y = (x + 1) * 4 + 1  →  -1→1, 0→5, 1→9
def joy_to_sam9(x):
    """Transforma señal joystick [-1,1] a escala SAM 1–9."""
    return (np.asarray(x, dtype=float) + 1.0) * 4.0 + 1.0

SCR_PEAK_THRESHOLD  = 0.05      # umbral mínimo de amplitud SCR (µS)
SMNA_PEAK_THRESHOLD = 0.5       # umbral mínimo de amplitud SMNA
SYMP_BAND           = (0.045, 0.25)   # Hz — banda simpática (Posada-Quintero et al., 2016)
ICC_GRID_POINTS     = 100       # puntos para remuestreo temporal del ICC

OUTPUT_DIR  = "results/descriptive"
FIGURES_DIR = os.path.join(OUTPUT_DIR, "figures")
os.makedirs(FIGURES_DIR, exist_ok=True)

VALIDATION_LOG_PATH         = "results/physio_validation_log.json"
MARKERS_VALIDATION_LOG_PATH = "results/markers_validation_log.json"

# Clasificación teórica → cuadrante
STIM_QUADRANT = {
    **{s: "HVLA" for s in [1, 5, 6]},
    **{s: "HVHA" for s in [2, 3, 4, 7]},
    **{s: "LVLA" for s in [8, 11, 13]},
    **{s: "LVHA" for s in [9, 10, 12, 14]},
}

QUAD_ORDER  = ["HVHA", "HVLA", "LVHA", "LVLA"]
QUAD_LABELS = {
    "HVHA": "Alta Valencia – Alto Arousal",
    "HVLA": "Alta Valencia – Bajo Arousal",
    "LVHA": "Baja Valencia – Alto Arousal",
    "LVLA": "Baja Valencia – Bajo Arousal",
}
QUAD_COLORS = {
    "HVHA": "#E8A838",
    "HVLA": "#5DADE2",
    "LVHA": "#E74C3C",
    "LVLA": "#82E0AA",
}

# Paleta de tesis: colores de encabezado por cuadrante (hex sin #)
QUAD_HEADER_HEX = {
    "HVHA": "F0C050",
    "HVLA": "ADD8E6",
    "LVHA": "F1948A",
    "LVLA": "A9DFBF",
}

matplotlib.rcParams.update({
    "font.family": "DejaVu Sans", "font.size": 11,
    "axes.spines.top": False, "axes.spines.right": False,
    "figure.dpi": 300,
})


# ─────────────────────────────────────────────────────────────────────────────
# UTILIDADES GENERALES
# ─────────────────────────────────────────────────────────────────────────────
def get_col(df, names):
    """Busca columnas por nombre, insensible a mayúsculas."""
    for n in names:
        if n in df.columns:
            return n
        for c in df.columns:
            if c.lower() == n.lower():
                return c
    return None


def robust_zscore(series):
    """Z-score robusto; devuelve ceros si la DE es 0."""
    m, s = series.mean(), series.std()
    return (series - m) / s if s > 0 else pd.Series(0.0, index=series.index)

def compute_session_stats(eda_filepaths):
    col_names = {}
    all_data  = {k: [] for k in ['clean', 'phasic', 'tonic', 'smna']}

    for fp in eda_filepaths:
        if not os.path.exists(fp):
            continue
        df_tmp = pd.read_csv(fp, sep='\t')
        if not col_names:
            col_names['clean']  = get_col(df_tmp, ['EDA_Clean', 'Clean'])
            col_names['phasic'] = get_col(df_tmp, ['EDA_Phasic', 'Phasic'])
            col_names['tonic']  = get_col(df_tmp, ['EDA_Tonic',  'Tonic'])
            col_names['smna']   = get_col(df_tmp, ['SMNA',       'smna'])
        for key, col in col_names.items():
            if col and col in df_tmp.columns:
                all_data[key].append(df_tmp[col].values)

    session_stats = {}
    for key, col in col_names.items():
        if all_data[key]:
            concat = np.concatenate(all_data[key])
            mean_val, std_val = np.nanmean(concat), np.nanstd(concat)
            session_stats[col] = (mean_val, std_val if std_val > 0 else 1.0)
        else:
            session_stats[col] = (0.0, 1.0)

    return session_stats, col_names

def spectral_power_band(signal, sfreq, band):
    """Potencia espectral integrada en una banda de frecuencia."""
    if len(signal) < 4:
        return np.nan
    nperseg = min(int(sfreq * 4), len(signal))
    freqs, psd = welch(signal, fs=sfreq, nperseg=nperseg)
    idx = (freqs >= band[0]) & (freqs <= band[1])
    if idx.sum() == 0:
        return np.nan
    return float(np.trapz(psd[idx], freqs[idx]))


def load_json(path):
    if os.path.exists(path):
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    return {}


# ─────────────────────────────────────────────────────────────────────────────
# ICC TEMPORAL
# ─────────────────────────────────────────────────────────────────────────────
def icc_two_way_consistency(matrix):
    """
    ICC(2,1) consistencia sobre matriz (sujetos × timepoints).
    Modelo two-way mixed (cada columna = timepoint compartido).
    Retorna ICC en [−1, 1].
    """
    matrix = np.array(matrix, dtype=float)
    n, k = matrix.shape   # n sujetos, k timepoints
    if n < 2 or k < 2:
        return np.nan

    grand_mean = np.nanmean(matrix)
    # SS between subjects (filas)
    row_means = np.nanmean(matrix, axis=1)
    SS_r = k * np.nansum((row_means - grand_mean) ** 2)
    # SS between timepoints (columnas)
    col_means = np.nanmean(matrix, axis=0)
    SS_c = n * np.nansum((col_means - grand_mean) ** 2)
    # SS error
    SS_tot = np.nansum((matrix - grand_mean) ** 2)
    SS_e = SS_tot - SS_r - SS_c

    df_r = n - 1
    df_c = k - 1
    df_e = df_r * df_c

    if df_e == 0:
        return np.nan

    MS_r = SS_r / df_r
    MS_e = SS_e / df_e

    icc = (MS_r - MS_e) / (MS_r + (k - 1) * MS_e)
    return float(np.clip(icc, -1, 1))


def spearman_brown_correction(icc, k_ref=60, k_actual=None):
    """
    Corrección Spearman-Brown: ajusta el ICC observado al que se esperaría
    con k_ref timepoints, dado que el estímulo tiene k_actual timepoints.
    Esto "deconfoundea" la duración: estímulos más largos tienen más timepoints
    y, en igualdad de señal, producen ICCs más altos.
    icc_adj = (k_ref/k_actual * icc) / (1 + (k_ref/k_actual - 1) * icc)
    """
    if k_actual is None or k_actual == 0 or icc is None or np.isnan(icc):
        return np.nan
    ratio = k_ref / k_actual
    denom = 1 + (ratio - 1) * icc
    if denom == 0:
        return np.nan
    return float(np.clip((ratio * icc) / denom, -1, 1))


def compute_temporal_icc(profiles_dict, n_points=ICC_GRID_POINTS):
    """
    Dado un dict {subject_id: array_1d}, remuestrea cada perfil a n_points
    en [0,1] y calcula ICC(2,1) de consistencia.

    Retorna:
      icc_raw      : ICC sin corregir
      icc_sb       : ICC corregido por Spearman-Brown (referencia = n_points)
      n_subjects   : N de sujetos con datos válidos
      mean_duration: duración media de los perfiles originales (en muestras)
    """
    grid = np.linspace(0, 1, n_points)
    rows = []
    raw_lengths = []

    for sid, profile in profiles_dict.items():
        profile = np.array(profile, dtype=float)
        if len(profile) < 4:
            continue
        raw_lengths.append(len(profile))
        x = np.linspace(0, 1, len(profile))
        try:
            f = interp1d(x, profile, kind="linear", fill_value="extrapolate")
            rows.append(f(grid))
        except Exception:
            continue

    if len(rows) < 2:
        return np.nan, np.nan, len(rows), np.nan

    matrix   = np.vstack(rows)  # (n_sujetos × n_points)
    icc_raw  = icc_two_way_consistency(matrix)
    icc_sb   = spearman_brown_correction(icc_raw, k_ref=n_points,
                                         k_actual=int(np.mean(raw_lengths)))
    return icc_raw, icc_sb, len(rows), float(np.mean(raw_lengths))


# ─────────────────────────────────────────────────────────────────────────────
# FORMATEADO EXCEL
# ─────────────────────────────────────────────────────────────────────────────
THIN  = Side(style="thin",   color="CCCCCC")
THICK = Side(style="medium", color="888888")

HEADER_FONT   = Font(bold=True, size=11, color="FFFFFF", name="Arial")
HEADER_FILL   = PatternFill("solid", start_color="2E4057")   # azul oscuro
SUBHEAD_FONT  = Font(bold=True, size=10, color="2E4057", name="Arial")
SUBHEAD_FILL  = PatternFill("solid", start_color="E8EDF2")
DATA_FONT     = Font(size=10, name="Arial")
ALT_FILL      = PatternFill("solid", start_color="F7F9FC")
BORDER_ALL    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_HEADER = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)


def _apply_header(ws, row, cols, fill=None):
    fill = fill or HEADER_FILL
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font   = HEADER_FONT
        cell.fill   = fill
        cell.border = BORDER_ALL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)


def _apply_data_row(ws, row, cols, alt=False):
    fill = ALT_FILL if alt else PatternFill("solid", start_color="FFFFFF")
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font   = DATA_FONT
        cell.fill   = fill
        cell.border = BORDER_ALL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autofit(ws, min_w=10, max_w=30):
    for col in ws.columns:
        length = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = \
            min(max(length + 2, min_w), max_w)


def _freeze(ws, cell="B2"):
    ws.freeze_panes = cell


def df_to_sheet(ws, df, title=None, quad_color_col=None):
    """
    Escribe un DataFrame en una hoja de openpyxl con formato tesis.
    quad_color_col: nombre de columna cuyo valor determina el color del encabezado.
    """
    start_row = 1
    if title:
        ws.merge_cells(start_row=1, start_column=1,
                       end_row=1, end_column=len(df.columns))
        tc = ws.cell(row=1, column=1)
        tc.value     = title
        tc.font      = Font(bold=True, size=12, color="FFFFFF", name="Arial")
        tc.fill      = PatternFill("solid", start_color="1A2F45")
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20
        start_row = 2

    # Encabezados de columna
    header_row = start_row
    quad_val   = None
    if quad_color_col and quad_color_col in df.columns:
        unique_quads = df[quad_color_col].dropna().unique()
        quad_val = unique_quads[0] if len(unique_quads) == 1 else None

    h_fill = HEADER_FILL
    if quad_val and quad_val in QUAD_HEADER_HEX:
        h_fill = PatternFill("solid", start_color=QUAD_HEADER_HEX[quad_val])
        h_font = Font(bold=True, size=11, color="1A2F45", name="Arial")
    else:
        h_font = HEADER_FONT

    for c_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=header_row, column=c_idx, value=col_name)
        cell.font      = h_font
        cell.fill      = h_fill
        cell.border    = BORDER_ALL
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[header_row].height = 28

    # Datos
    for r_idx, (_, row_data) in enumerate(df.iterrows(), 1):
        row_num = header_row + r_idx
        alt     = (r_idx % 2 == 0)
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=c_idx)
            cell.value     = value
            cell.font      = DATA_FONT
            cell.fill      = ALT_FILL if alt else PatternFill("solid",
                                                               start_color="FFFFFF")
            cell.border    = BORDER_ALL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Formato numérico
            if isinstance(value, float):
                cell.number_format = "0.0000"

    _autofit(ws)
    _freeze(ws)
    return ws


def save_workbook_with_sheets(filepath, sheets_dict):
    """
    sheets_dict: {sheet_name: (df, title, quad_color_col)}
    """
    wb = Workbook()
    wb.remove(wb.active)   # quitar hoja vacía default
    for sheet_name, (df, title, quad_col) in sheets_dict.items():
        ws = wb.create_sheet(title=sheet_name[:31])
        df_to_sheet(ws, df, title=title, quad_color_col=quad_col)
    wb.save(filepath)
    print(f"  ✅ Guardado: {filepath}")


# ─────────────────────────────────────────────────────────────────────────────
# ESTADÍSTICOS DESCRIPTIVOS
# ─────────────────────────────────────────────────────────────────────────────
def desc_stats_series(values):
    """Retorna dict con estadísticos de una serie numérica."""
    v = np.array(values, dtype=float)
    v = v[~np.isnan(v)]
    if len(v) == 0:
        return dict(n=0, mean=np.nan, median=np.nan, sd=np.nan,
                    sem=np.nan, min=np.nan, max=np.nan, cv=np.nan)
    sem = float(np.std(v, ddof=1) / np.sqrt(len(v))) if len(v) > 1 else np.nan
    sd  = float(np.std(v, ddof=1)) if len(v) > 1 else np.nan
    cv  = float(sd / np.abs(np.mean(v))) if np.mean(v) != 0 else np.nan
    return dict(
        n       = int(len(v)),
        mean    = round(float(np.mean(v)),   4),
        median  = round(float(np.median(v)), 4),
        sd      = round(sd,   4),
        sem     = round(sem,  4),
        min     = round(float(np.min(v)), 4),
        max     = round(float(np.max(v)), 4),
        cv      = round(cv,   4) if not np.isnan(cv) else np.nan,
    )


# ─────────────────────────────────────────────────────────────────────────────
# CARGA LOGS DE VALIDACIÓN
# ─────────────────────────────────────────────────────────────────────────────
validation_data         = load_json(VALIDATION_LOG_PATH).get("subjects", {})
markers_validation_data = load_json(MARKERS_VALIDATION_LOG_PATH).get("subjects", {})


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACCIÓN NIVEL 1: una fila por sujeto × estímulo
# ─────────────────────────────────────────────────────────────────────────────
# Además del registro estático, guardamos los perfiles temporales del joystick
# para calcular el ICC luego (no se persisten al CSV, solo en memoria).
records      = []                   # cada dict → una fila del CSV nivel 1
joy_profiles = {}                   # {(stimulus, dimension): {subject_id: array}}
raw_joy_profiles = {} 
eda_profiles = {}                   # <--- NUEVO: para guardar perfiles EDA

print(f"\n{'='*60}")
print(f"  EXTRACCIÓN — {len(SUBJECT_IDS)} sujetos")
print(f"{'='*60}\n")

for subject_id in SUBJECT_IDS:
    print(f"  Sujeto {subject_id}")

    base_path       = rf"data/derivatives/campeones_preproc/sub-{subject_id}/ses-vr/eeg"
    sourcedata_path = rf"data/sourcedata/xdf/sub-{subject_id}"
    eda_base_path   = rf"data/derivatives/eda_preproc_tests/sub-{subject_id}"

    if not os.path.exists(base_path):
        print(f"    ⚠ base_path no existe, saltando.")
        continue

    eeg_files = sorted([f for f in os.listdir(base_path)
                        if f.endswith("_desc-preproc_eeg.vhdr")])

    # ── CAMBIO 2: pre-computar stats de normalización por sesión ──────────
    bad_blocks = set()
    for eeg_filename in eeg_files:
        skip_block = False
        if subject_id in validation_data:
            for val_key, val_info in validation_data[subject_id].items():
                if val_key in eeg_filename:
                    cat = (val_info.get("eda") or val_info.get("gsr") or {}) \
                          .get("category", "good").lower()
                    if cat in ["bad", "maybe"]:
                        skip_block = True
                    break
        if subject_id in markers_validation_data and not skip_block:
            for val_key, note in markers_validation_data[subject_id].items():
                if val_key in eeg_filename:
                    nu = note.strip().upper()
                    is_bad_stim = bool(re.search(r"ESTIMULO BAD|BAD ESTIMULO", nu))
                    if not is_bad_stim and (nu.startswith("BAD") or
                       "NO SE TOMO" in nu or "NO TOMADO" in nu):
                        skip_block = True
                    break
        if skip_block:
            bad_blocks.add(eeg_filename)

    session_files_map = {}
    for eeg_filename in eeg_files:
        if eeg_filename in bad_blocks:
            continue
        parts   = eeg_filename.split('_')
        task_id = parts[2].split('-')[1]
        acq_id  = parts[3].split('-')[1].upper()
        eda_fp  = os.path.join(
            eda_base_path,
            f"sub-{subject_id}_ses-{acq_id}_task-{task_id}_desc-edapreproc_physio.tsv"
        )
        session_files_map.setdefault(acq_id, []).append(eda_fp)

    session_stats_cache = {}
    for acq_id, eda_files in session_files_map.items():
        stats, _ = compute_session_stats(eda_files)
        session_stats_cache[acq_id] = stats
    # ── FIN CAMBIO 2 ───────────────────────────────────────────────────────

    for eeg_filename in eeg_files:   # ← el loop original sigue igual desde acá
        skip_block  = False
        bad_stimuli = []

        # ── Validaciones de bloque ─────────────────────────────
        if subject_id in validation_data:
            for val_key, val_info in validation_data[subject_id].items():
                if val_key in eeg_filename:
                    cat = (val_info.get("eda") or val_info.get("gsr") or {}) \
                          .get("category", "good").lower()
                    if cat in ["bad", "maybe"]:
                        skip_block = True
                    break

        if subject_id in markers_validation_data and not skip_block:
            for val_key, note in markers_validation_data[subject_id].items():
                if val_key in eeg_filename:
                    nu = note.strip().upper()
                    if re.search(r"ESTIMULO BAD|BAD ESTIMULO", nu):
                        m = re.search(r'\[(.*?)\]', note)
                        if m:
                            bad_stimuli.append(m.group(1).strip().lower())
                    elif (nu.startswith("BAD") or "NO SE TOMO" in nu
                          or "NO TOMADO" in nu):
                        skip_block = True
                    break

        if skip_block:
            continue

        # ── Identificación de bloque ───────────────────────────
        parts   = eeg_filename.split("_")
        task_id = parts[2].split("-")[1]
        acq_id  = parts[3].split("-")[1].upper()

        # ── Cargar EDA ─────────────────────────────────────────
        eda_fp = os.path.join(
            eda_base_path,
            f"sub-{subject_id}_ses-{acq_id}_task-{task_id}_desc-edapreproc_physio.tsv",
        )
        if not os.path.exists(eda_fp):
            continue

        eda_df = pd.read_csv(eda_fp, sep="\t")
        col_time   = get_col(eda_df, ["Time",      "time"])
        col_clean  = get_col(eda_df, ["EDA_Clean", "Clean"])
        col_phasic = get_col(eda_df, ["EDA_Phasic","Phasic"])
        col_tonic  = get_col(eda_df, ["EDA_Tonic", "Tonic"])
        col_smna   = get_col(eda_df, ["SMNA",      "smna"])

        if any(c is None for c in [col_time, col_clean, col_phasic,
                                    col_tonic, col_smna]):
            continue

        eda_sfreq = 1.0 / np.nanmean(np.diff(eda_df[col_time].values))

        # z-score a nivel de sesión
        session_stats = session_stats_cache.get(acq_id, {})
        eda_z = eda_df.copy()
        for col in [col_clean, col_phasic, col_tonic, col_smna]:
            mean_val, std_val = session_stats.get(col, (0.0, 1.0))
            eda_z[col] = (eda_df[col] - mean_val) / std_val

        # ── Cargar EEG (joystick) ──────────────────────────────
        try:
            raw = mne.io.read_raw_brainvision(
                os.path.join(base_path, eeg_filename),
                preload=True, verbose=False)
        except Exception as e:
            print(f"    ⚠ Error cargando EEG: {e}")
            continue

        eeg_sfreq  = raw.info["sfreq"]
        joy_data,_ = raw.get_data(picks=[CHANNEL_TO_PLOT], return_times=True)
        raw_joy    = joy_data[0].copy()

        # ── Diseño experimental ────────────────────────────────
        excel_path = os.path.join(
            sourcedata_path,
            f"order_matrix_{subject_id}_{acq_id}_block{int(task_id)}_VR.xlsx",
        )
        try:
            excel_df  = pd.read_excel(excel_path)
            dim       = ("valence"
                         if "valence" in excel_df["dimension"].dropna().tolist()
                         else "arousal")
            video_ids = excel_df["video_id"].dropna().tolist()
            inv_instr = (excel_df["order_emojis_slider"].dropna().iloc[0]
                         if "order_emojis_slider" in excel_df.columns else None)
        except Exception:
            continue

        # ── Eventos ────────────────────────────────────────────
        events_path = os.path.join(
            base_path, eeg_filename.replace("_eeg.vhdr", "_events.tsv"))
        if not os.path.exists(events_path):
            continue

        events_df     = pd.read_csv(events_path, sep="\t")
        video_counter = 0

        for _, row in events_df.iterrows():
            if row["trial_type"] != "video":
                continue
            if video_counter >= len(video_ids):
                break

            vid_str = str(video_ids[video_counter]).strip().lower()
            try:
                vid_str = str(int(float(vid_str)))
            except Exception:
                pass
            video_counter += 1

            if vid_str in bad_stimuli:
                continue

            try:
                vid_id = int(vid_str)
            except Exception:
                continue

            quadrant = STIM_QUADRANT.get(vid_id)
            if quadrant is None:
                continue

            onset  = float(row["onset"])
            offset = onset + float(row["duration"])

            # ══════════════════════════════════════════════════
            # AUTORREPORTE — señal continua completa
            # ══════════════════════════════════════════════════
            s0  = int(onset  * eeg_sfreq)
            s1  = int(offset * eeg_sfreq)
            beh = raw_joy[s0:s1].copy()

            # ... (líneas existentes)
            s1  = int(offset * eeg_sfreq)
            beh = raw_joy[s0:s1].copy()

            # 1) Corregir inversión de instrucción (si aplica)
            if inv_instr in ["inverse", "indirect"]:
                beh *= -1

            # 2) Transformar a escala SAM 1–9 ANTES de cualquier agregación
            beh = joy_to_sam9(beh)

            # --- Guardar perfil transformado [1, 9] para los plots 2D ---
            key_raw = (vid_id, dim)
            if key_raw not in raw_joy_profiles:
                raw_joy_profiles[key_raw] = {}
            raw_joy_profiles[key_raw][subject_id] = beh.copy()

            # 3) Suavizado
            win    = max(1, int(SMOOTHING_WINDOW_S * eeg_sfreq))
            beh_sm = (pd.Series(beh)
                      .rolling(win, center=True, min_periods=1)
                      .mean().values)

            # Métricas de autorreporte (todas en escala 1–9)
            rep_mean   = float(np.mean(beh_sm))
            rep_median = float(np.median(beh_sm))
            rep_sd     = float(np.std(beh_sm, ddof=1)) if len(beh_sm) > 1 else np.nan
            rep_min    = float(np.min(beh_sm))
            rep_max    = float(np.max(beh_sm))
            rep_range  = rep_max - rep_min

            # Derivada media: velocidad de cambio (unidades SAM/s)
            rep_deriv_mean = float(
                np.mean(np.abs(np.gradient(beh_sm) * eeg_sfreq))
            )

            # Guardar perfil temporal para ICC (en escala 1–9)
            key = (vid_id, dim)
            if key not in joy_profiles:
                joy_profiles[key] = {}
            joy_profiles[key][subject_id] = beh_sm

            # ══════════════════════════════════════════════════
            # EDA — señal completa del estímulo
            # ══════════════════════════════════════════════════
            mask       = ((eda_df[col_time] >= onset) &
                          (eda_df[col_time] <  offset))
            eda_stim   = eda_df[mask].copy()
            eda_stim_z = eda_z[mask].copy()

            if len(eda_stim) < 4:
                continue

            t_vals     = eda_stim[col_time].values
            clean_raw  = eda_stim[col_clean].values
            phasic_raw = eda_stim[col_phasic].values
            tonic_raw  = eda_stim[col_tonic].values
            smna_raw   = eda_stim[col_smna].values
            phasic_z   = eda_stim_z[col_phasic].values
            tonic_z    = eda_stim_z[col_tonic].values
            smna_z     = eda_stim_z[col_smna].values

            session_ab = acq_id[-1].upper()  # Extraemos 'A' o 'B' de la sesión
            key_eda = (vid_id, session_ab)
            if key_eda not in eda_profiles:
                eda_profiles[key_eda] = {}
            eda_profiles[key_eda][subject_id] = eda_stim_z[col_clean].values.copy()

            # ── Tónico y fásico ──────────────────────────────
            tonic_mean  = float(np.nanmean(tonic_z))
            tonic_med   = float(np.nanmedian(tonic_z))
            tonic_sd    = (float(np.nanstd(tonic_z, ddof=1))
                           if np.sum(~np.isnan(tonic_z)) > 1 else np.nan)
            phasic_mean = float(np.nanmean(phasic_z))
            phasic_med  = float(np.nanmedian(phasic_z))
            phasic_sd   = (float(np.nanstd(phasic_z, ddof=1))
                           if np.sum(~np.isnan(phasic_z)) > 1 else np.nan)

            # ── SCR ──────────────────────────────────────────
            p_nn = np.nan_to_num(phasic_raw)
            try:
                _, scr_info   = nk.eda_peaks(p_nn, sampling_rate=eda_sfreq,
                                              amplitude_min=SCR_PEAK_THRESHOLD)
                scr_count     = int(len(scr_info["SCR_Peaks"]))
                scr_amp_mean  = (float(np.nanmean(scr_info["SCR_Amplitude"]))
                                 if scr_count > 0 else 0.0)
                scr_rise_mean = (float(np.nanmean(scr_info["SCR_RiseTime"]))
                                 if scr_count > 0 else 0.0)
            except Exception:
                scr_count, scr_amp_mean, scr_rise_mean = 0, 0.0, 0.0

            dur_min          = (offset - onset) / 60.0
            scr_rate_per_min = scr_count / dur_min if dur_min > 0 else np.nan

            # ── SMNA ─────────────────────────────────────────
            s_nn     = np.nan_to_num(smna_raw)
            dist_smp = int(eda_sfreq * 1.5)
            smna_peaks, smna_props = find_peaks(
                s_nn, height=SMNA_PEAK_THRESHOLD, distance=dist_smp)
            smna_count    = int(len(smna_peaks))
            smna_amp_mean = (float(np.nanmean(smna_props["peak_heights"]))
                             if smna_count > 0 else 0.0)
            valid_smna    = ~np.isnan(smna_raw) # <--- Usar señal sin z-score
            smna_auc      = (float(np.trapz(smna_raw[valid_smna], x=t_vals[valid_smna])) if np.sum(valid_smna) > 1 else 0.0)
            
            # ── Potencia banda simpática ─────────────────────
            sym_power = spectral_power_band(
                np.nan_to_num(phasic_raw), eda_sfreq, SYMP_BAND)

            # ── Variabilidad EDA ──────────────────────────────
            eda_range    = float(np.nanmax(clean_raw) - np.nanmin(clean_raw))
            eda_clean_sd = (float(np.nanstd(clean_raw, ddof=1))
                            if np.sum(~np.isnan(clean_raw)) > 1 else np.nan)

            records.append(dict(
                subject_id=subject_id, session=acq_id,
                block=task_id, stimulus=vid_id,
                quadrant=quadrant, dimension=dim,
                duration_s=round(offset - onset, 2),
                # Autorreporte (escala SAM 1–9)
                rep_mean=rep_mean,
                rep_median=rep_median, rep_sd=rep_sd,
                rep_min=rep_min, rep_max=rep_max,
                rep_range=rep_range, rep_deriv_mean=rep_deriv_mean,
                # EDA tónico/fásico
                tonic_mean=tonic_mean, tonic_med=tonic_med, tonic_sd=tonic_sd,
                phasic_mean=phasic_mean, phasic_med=phasic_med,
                phasic_sd=phasic_sd,
                # SCR
                scr_count=scr_count, scr_amp_mean=scr_amp_mean,
                scr_rise_mean=scr_rise_mean, scr_rate_per_min=scr_rate_per_min,
                # SMNA
                smna_count=smna_count, smna_amp_mean=smna_amp_mean,
                smna_auc=smna_auc,
                # Espectral
                sym_band_power=sym_power,
                # Variabilidad EDA
                eda_range=eda_range, eda_clean_sd=eda_clean_sd,
            ))

print(f"\nRegistros extraídos (sujeto × estímulo): {len(records)}")
if not records:
    raise RuntimeError("No se obtuvieron registros. Verificar rutas y archivos.")

df = pd.DataFrame(records)

features_to_normalize = ['smna_auc', 'sym_band_power', 'scr_amp_mean']

for feat in features_to_normalize:
    if feat in df.columns:
        # Agrupamos explícitamente por sujeto Y sesión
        df[feat] = df.groupby(['subject_id', 'session'])[feat].transform(robust_zscore)

print("✅ Features escalares (AUC, Potencia, Amplitud SCR) normalizadas por sujeto y sesión.")

# ── Guardar nivel 1 como CSV ───────────────────────────────────────────────
csv_path = os.path.join(OUTPUT_DIR, "stimulus_level_data.csv")
df.to_csv(csv_path, index=False, encoding="utf-8-sig")
print(f"✅ Nivel 1 guardado: {csv_path}\n")


# ─────────────────────────────────────────────────────────────────────────────
# NIVEL 2: AGREGADO POR ESTÍMULO
# Primero calcular ICC temporal por estímulo, luego agregar métricas.
# ─────────────────────────────────────────────────────────────────────────────
print("Calculando ICC temporal por estímulo...")

icc_records = []
for (vid_id, dim), profiles in joy_profiles.items():
    icc_raw, icc_sb, n_subj, mean_len = compute_temporal_icc(profiles)
    icc_records.append(dict(
        stimulus=vid_id, dimension=dim,
        icc_raw=icc_raw, icc_sb=icc_sb,
        icc_n_subjects=n_subj,
        icc_mean_profile_length=mean_len,
    ))

icc_df = pd.DataFrame(icc_records)


def build_stimulus_agg(df_in, dim_filter, rep_col, rep_label,
                       extra_rep_cols=None):
    """
    Agrega el DataFrame nivel 1 por estímulo.
    Para cada métrica calcula: media, mediana, SD, SEM, min, max entre sujetos.
    También incluye ICC temporal y CV del estímulo.

    rep_col: columna principal de autorreporte
    extra_rep_cols: lista de (col, label) adicionales
    """
    sub = df_in[df_in["dimension"] == dim_filter].copy()

    # Columnas EDA a agregar
    eda_cols = [
        ("tonic_mean",       "Tónico Media (z)"),
        ("tonic_sd",         "Tónico SD (z)"),
        ("phasic_mean",      "Fásico Media (z)"),
        ("phasic_sd",        "Fásico SD (z)"),
        ("scr_amp_mean",     "Amplitud SCR (z)"),
        ("scr_rise_mean",    "Tiempo Subida SCR (s)"),
        ("scr_rate_per_min", "Frecuencia SCR (peaks/min)"),
        ("smna_auc",         "SMNA AUC (z·s)"),
        ("sym_band_power",   "Potencia Simpática (z²/Hz)"),
        ("eda_range",        "Rango EDA (µS)"),
        ("eda_clean_sd",     "EDA SD (µS)"),
    ]

    rows = []
    for stim_id in sorted(sub["stimulus"].unique()):
        s = sub[sub["stimulus"] == stim_id]
        q = STIM_QUADRANT.get(stim_id, "?")
        dur_vals = s["duration_s"].values

        # ICC
        icc_row = icc_df[
            (icc_df["stimulus"]  == stim_id) &
            (icc_df["dimension"] == dim_filter)
        ]
        icc_raw = icc_row["icc_raw"].values[0] if len(icc_row) else np.nan
        icc_sb  = icc_row["icc_sb"].values[0]  if len(icc_row) else np.nan
        n_icc   = int(icc_row["icc_n_subjects"].values[0]) if len(icc_row) else 0

        # CV del estímulo (sobre rep_col): cuánta variabilidad intrínseca tiene
        rep_vals = s[rep_col].dropna().values
        stim_cv  = (float(np.std(rep_vals, ddof=1) / np.abs(np.mean(rep_vals)))
                    if len(rep_vals) > 1 and np.mean(rep_vals) != 0 else np.nan)

        rec = {
            "Estímulo":          stim_id,
            "Cuadrante":         q,
            "Cuadrante (label)": QUAD_LABELS.get(q, q),
            "N sujetos":         int(len(s)),
            "Duración media (s)":round(float(np.mean(dur_vals)), 2),
            "Duración SD (s)":   round(float(np.std(dur_vals, ddof=1)), 2)
                                  if len(dur_vals) > 1 else np.nan,
        }

        # Autorreporte principal (escala SAM 1–9)
        ds = desc_stats_series(rep_vals)
        for k, v in ds.items():
            if k == "n":
                continue
            rec[f"{rep_label} – {k}"] = v

        # Intra-sujeto SD del reporte continuo (promedio entre sujetos)
        intra_sd_vals = s["rep_sd"].dropna().values
        rec[f"{rep_label} – SD intra-sujeto (media)"] = (
            round(float(np.mean(intra_sd_vals)), 4) if len(intra_sd_vals) else np.nan)

        # Rango intra-sujeto
        rep_range_vals = s["rep_range"].dropna().values
        rec[f"{rep_label} – Rango intra-sujeto (media)"] = (
            round(float(np.mean(rep_range_vals)), 4) if len(rep_range_vals) else np.nan)

        # Derivada media
        deriv_vals = s["rep_deriv_mean"].dropna().values
        rec[f"{rep_label} – Deriv. media (media)"] = (
            round(float(np.mean(deriv_vals)), 4) if len(deriv_vals) else np.nan)

        # CV
        rec[f"{rep_label} – CV entre sujetos"] = (
            round(stim_cv, 4) if not np.isnan(stim_cv) else np.nan)

        # ICC temporal
        rec["ICC temporal (raw)"]     = round(icc_raw, 4) if not np.isnan(icc_raw) else np.nan
        rec["ICC temporal (SB-adj)"]  = round(icc_sb,  4) if not np.isnan(icc_sb)  else np.nan
        rec["ICC N sujetos"]          = n_icc
        rec["Nota SB-adj"] = (
            "ICC ajustado por duración (Spearman-Brown, referencia = 100 puntos). "
            "CV entre sujetos es covariable interpretativa: "
            "mayor CV → ICC bajo esperable."
        )

        # EDA
        for col, lbl in eda_cols:
            if col in s.columns:
                ds_e = desc_stats_series(s[col].dropna().values)
                for k, v in ds_e.items():
                    if k == "n":
                        continue
                    rec[f"{lbl} – {k}"] = v

        rows.append(rec)

    return pd.DataFrame(rows)


print("Construyendo tablas de nivel 2 (por estímulo)...")
val_df = df[df["dimension"] == "valence"]
aro_df = df[df["dimension"] == "arousal"]

stim_agg_val = build_stimulus_agg(df, "valence", "rep_mean",
                                   "Valencia (escala 1–9)")
stim_agg_aro = build_stimulus_agg(df, "arousal", "rep_mean",
                                   "Arousal (escala 1–9)")
stim_agg_eda = build_stimulus_agg(df, "valence", "rep_mean",
                                   "Valencia")   # EDA no depende de dimensión
# Para EDA usamos todos los registros; el agg de arousal ya tiene EDA también.
# Generamos una tabla EDA pura:
def build_eda_stimulus_agg(df_in):
    """Agrega solo features EDA por estímulo (independiente de dimensión)."""
    eda_cols = [
        ("tonic_mean",       "Tónico Media (z)"),
        ("tonic_sd",         "Tónico SD (z)"),
        ("phasic_mean",      "Fásico Media (z)"),
        ("phasic_sd",        "Fásico SD (z)"),
        ("scr_amp_mean",     "Amplitud SCR (z)"),
        ("scr_rise_mean",    "Tiempo Subida SCR (s)"),
        ("scr_rate_per_min", "Frecuencia SCR (peaks/min)"),
        ("smna_auc",         "SMNA AUC (z·s)"),
        ("sym_band_power",   "Potencia Simpática (z²/Hz)"),
        ("eda_range",        "Rango EDA (µS)"),
        ("eda_clean_sd",     "EDA SD (µS)"),
    ]
    rows = []
    for stim_id in sorted(df_in["stimulus"].unique()):
        s = df_in[df_in["stimulus"] == stim_id]
        q = STIM_QUADRANT.get(stim_id, "?")
        rec = {
            "Estímulo":          stim_id,
            "Cuadrante":         q,
            "Cuadrante (label)": QUAD_LABELS.get(q, q),
            "N registros":       int(len(s)),
        }
        for col, lbl in eda_cols:
            if col not in s.columns:
                continue
            ds = desc_stats_series(s[col].dropna().values)
            for k, v in ds.items():
                if k == "n":
                    continue
                rec[f"{lbl} – {k}"] = v
        rows.append(rec)
    return pd.DataFrame(rows)


stim_agg_eda_pure = build_eda_stimulus_agg(df)

# Guardar nivel 2
save_workbook_with_sheets(
    os.path.join(OUTPUT_DIR, "stimulus_agg_valence.xlsx"),
    {"Valencia por estímulo": (stim_agg_val, "Nivel 2: Autorreporte Valencia — Agregado por Estímulo", "Cuadrante")},
)
save_workbook_with_sheets(
    os.path.join(OUTPUT_DIR, "stimulus_agg_arousal.xlsx"),
    {"Arousal por estímulo": (stim_agg_aro, "Nivel 2: Autorreporte Arousal — Agregado por Estímulo", "Cuadrante")},
)
save_workbook_with_sheets(
    os.path.join(OUTPUT_DIR, "stimulus_agg_eda.xlsx"),
    {"EDA por estímulo": (stim_agg_eda_pure, "Nivel 2: EDA — Agregado por Estímulo", "Cuadrante")},
)


# ─────────────────────────────────────────────────────────────────────────────
# NIVEL 3: AGREGADO POR CUADRANTE
# Se parte del nivel 2 (valores por estímulo) y se agrega por cuadrante.
# ─────────────────────────────────────────────────────────────────────────────
def build_quadrant_agg(stim_df, title_prefix):
    """
    Toma el DataFrame de nivel 2 (una fila por estímulo) y colapsa por cuadrante.
    Para cada columna numérica calcula media, mediana, SD entre estímulos.
    """
    numeric_cols = stim_df.select_dtypes(include=[np.number]).columns.tolist()
    # Excluir columnas de ID
    exclude = {"Estímulo", "N sujetos", "N registros", "ICC N sujetos"}
    numeric_cols = [c for c in numeric_cols if c not in exclude]

    rows = []
    for q in QUAD_ORDER:
        sub = stim_df[stim_df["Cuadrante"] == q]
        if len(sub) == 0:
            continue
        rec = {
            "Cuadrante":         q,
            "Cuadrante (label)": QUAD_LABELS[q],
            "N estímulos":       int(len(sub)),
            "N sujetos (media)": round(sub["N sujetos"].mean(), 1)
                                  if "N sujetos" in sub.columns else np.nan,
        }
        for col in numeric_cols:
            vals = sub[col].dropna().values
            if len(vals) == 0:
                rec[f"{col} – media"]   = np.nan
                rec[f"{col} – mediana"] = np.nan
                rec[f"{col} – SD"]      = np.nan
            else:
                rec[f"{col} – media"]   = round(float(np.mean(vals)),   4)
                rec[f"{col} – mediana"] = round(float(np.median(vals)), 4)
                rec[f"{col} – SD"]      = (round(float(np.std(vals, ddof=1)), 4)
                                            if len(vals) > 1 else np.nan)
        rows.append(rec)
    return pd.DataFrame(rows)


print("Construyendo tablas de nivel 3 (por cuadrante)...")
quad_val = build_quadrant_agg(stim_agg_val,      "Valencia")
quad_aro = build_quadrant_agg(stim_agg_aro,      "Arousal")
quad_eda = build_quadrant_agg(stim_agg_eda_pure, "EDA")

# Guardar nivel 3
save_workbook_with_sheets(
    os.path.join(OUTPUT_DIR, "quadrant_desc_valence.xlsx"),
    {"Valencia por cuadrante": (quad_val,
                                "Nivel 3: Autorreporte Valencia — Descriptivo por Cuadrante",
                                "Cuadrante")},
)
save_workbook_with_sheets(
    os.path.join(OUTPUT_DIR, "quadrant_desc_arousal.xlsx"),
    {"Arousal por cuadrante": (quad_aro,
                               "Nivel 3: Autorreporte Arousal — Descriptivo por Cuadrante",
                               "Cuadrante")},
)
save_workbook_with_sheets(
    os.path.join(OUTPUT_DIR, "quadrant_desc_eda.xlsx"),
    {"EDA por cuadrante": (quad_eda,
                           "Nivel 3: EDA — Descriptivo por Cuadrante",
                           "Cuadrante")},
)


# ─────────────────────────────────────────────────────────────────────────────
# OVERVIEW DEL DATASET
# ─────────────────────────────────────────────────────────────────────────────
print("Construyendo overview del dataset...")

# Tabla 1: N de registros por sujeto
n_per_subj = (df.groupby("subject_id")
                .agg(
                    n_registros=("stimulus","count"),
                    n_estimulos_unicos=("stimulus","nunique"),
                    dim_valence=("dimension", lambda x: (x=="valence").sum()),
                    dim_arousal=("dimension", lambda x: (x=="arousal").sum()),
                    dur_media=("duration_s", "mean"),
                    dur_sd=("duration_s", "std"),
                )
                .reset_index()
                .rename(columns={
                    "subject_id":        "Sujeto",
                    "n_registros":       "N registros",
                    "n_estimulos_unicos":"N estímulos únicos",
                    "dim_valence":       "N bloques Valencia",
                    "dim_arousal":       "N bloques Arousal",
                    "dur_media":         "Duración media (s)",
                    "dur_sd":            "Duración SD (s)",
                }))
for c in ["Duración media (s)", "Duración SD (s)"]:
    n_per_subj[c] = n_per_subj[c].round(2)

# Tabla 2: N registros por cuadrante y dimensión
n_per_quad = (df.groupby(["quadrant","dimension"])
                .size().unstack(fill_value=0)
                .reset_index()
                .rename(columns={"quadrant":"Cuadrante",
                                  "valence": "N Valencia",
                                  "arousal": "N Arousal"}))
n_per_quad.insert(1, "Cuadrante (label)",
                  n_per_quad["Cuadrante"].map(QUAD_LABELS))

# Tabla 3: Estadísticos de duración por estímulo
dur_per_stim = (df.groupby(["stimulus","quadrant"])["duration_s"]
                  .agg(["mean","median","std","min","max","count"])
                  .reset_index()
                  .rename(columns={
                      "stimulus": "Estímulo", "quadrant": "Cuadrante",
                      "mean": "Media (s)", "median": "Mediana (s)",
                      "std":  "SD (s)",    "min":   "Mín (s)",
                      "max":  "Máx (s)",   "count": "N sujetos",
                  }))
dur_per_stim.insert(2, "Cuadrante (label)",
                    dur_per_stim["Cuadrante"].map(QUAD_LABELS))
for c in ["Media (s)","Mediana (s)","SD (s)","Mín (s)","Máx (s)"]:
    dur_per_stim[c] = dur_per_stim[c].round(2)

# Tabla 4: ICC resumen
icc_summary = icc_df.copy()
icc_summary["Cuadrante"] = icc_summary["stimulus"].map(STIM_QUADRANT)
icc_summary["Cuadrante (label)"] = icc_summary["Cuadrante"].map(QUAD_LABELS)
icc_summary = icc_summary.rename(columns={
    "stimulus":               "Estímulo",
    "dimension":              "Dimensión",
    "icc_raw":                "ICC temporal (raw)",
    "icc_sb":                 "ICC temporal (SB-adj)",
    "icc_n_subjects":         "N sujetos",
    "icc_mean_profile_length":"Long. media perfil (muestras)",
})
icc_summary = icc_summary[[
    "Estímulo","Cuadrante","Cuadrante (label)","Dimensión",
    "N sujetos","Long. media perfil (muestras)",
    "ICC temporal (raw)","ICC temporal (SB-adj)",
]]
for c in ["ICC temporal (raw)","ICC temporal (SB-adj)"]:
    icc_summary[c] = icc_summary[c].round(4)

save_workbook_with_sheets(
    os.path.join(OUTPUT_DIR, "dataset_overview.xlsx"),
    {
        "N por sujeto":   (n_per_subj,   "Registros por Sujeto",                   None),
        "N por cuadrante":(n_per_quad,   "Registros por Cuadrante y Dimensión",    "Cuadrante"),
        "Duraciones":     (dur_per_stim, "Duración de Estímulos por Sujeto",       "Cuadrante"),
        "ICC":            (icc_summary,  "ICC Temporal por Estímulo",              "Cuadrante"),
    },
)


# ─────────────────────────────────────────────────────────────────────────────
# GRÁFICOS (300 dpi, listos para tesis)
# ─────────────────────────────────────────────────────────────────────────────
print("\nGenerando figuras...")

quad_xlabels     = [QUAD_LABELS[q].replace(" – ","\n") for q in QUAD_ORDER]
quad_colors_list = [QUAD_COLORS[q] for q in QUAD_ORDER]


def boxplot_panel(ax, data, col, title, ylabel, sig_pairs=None):
    """Boxplot con jitter por cuadrante."""
    groups = [data[data["quadrant"] == q][col].dropna().values for q in QUAD_ORDER]
    bp = ax.boxplot(groups, patch_artist=True,
                    medianprops=dict(color="black", linewidth=2),
                    whiskerprops=dict(linewidth=1.2),
                    capprops=dict(linewidth=1.2),
                    flierprops=dict(marker="o", markersize=3, alpha=0.4))
    for patch, q in zip(bp["boxes"], QUAD_ORDER):
        patch.set_facecolor(QUAD_COLORS[q])
        patch.set_alpha(0.80)
    for i, (g, q) in enumerate(zip(groups, QUAD_ORDER), 1):
        jit = np.random.normal(i, 0.07, size=len(g))
        ax.scatter(jit, g, color=QUAD_COLORS[q], alpha=0.4, s=14, zorder=3)
    ax.set_xticks(range(1, len(QUAD_ORDER)+1))
    ax.set_xticklabels(quad_xlabels, fontsize=9)
    ax.set_title(title, fontsize=11, fontweight="bold", pad=6)
    ax.set_ylabel(ylabel, fontsize=10)



def add_legend(fig):
    patches = [mpatches.Patch(color=QUAD_COLORS[q], label=QUAD_LABELS[q])
               for q in QUAD_ORDER]
    fig.legend(handles=patches, loc="lower center", ncol=2,
               fontsize=9, frameon=False,
               bbox_to_anchor=(0.5, -0.04))


# ── Fig 1: Autorreporte Valencia ──────────────────────────────────────────
fig, axes = plt.subplots(1, 2, figsize=(13, 5.5))
boxplot_panel(axes[0], val_df, "rep_mean",
              "Media del reporte", "Reporte promedio")
axes[0].axhline(5, color="gray", linewidth=0.8, linestyle="--", alpha=0.6)
boxplot_panel(axes[1], val_df, "rep_sd",
              "Variabilidad intra-sujeto",
              "DE")
plt.tight_layout(rect=[0, 0.05, 1, 1])
fig.savefig(os.path.join(FIGURES_DIR, "fig01_valencia_distribution.png"),
            dpi=300, bbox_inches="tight")
plt.close(fig)
print("  → fig01 guardada")

# ── Fig 2: Autorreporte Arousal ───────────────────────────────────────────
fig, axes = plt.subplots(1, 2, figsize=(13, 5.5))
boxplot_panel(axes[0], aro_df, "rep_mean",
              "Media del reporte", "Reporte promedio")
axes[0].axhline(5, color="gray", linewidth=0.8, linestyle="--", alpha=0.6)
boxplot_panel(axes[1], aro_df, "rep_sd",
              "Variabilidad intra-sujeto",
              "DE")
plt.tight_layout(rect=[0, 0.05, 1, 1])
fig.savefig(os.path.join(FIGURES_DIR, "fig02_arousal_distribution.png"),
            dpi=300, bbox_inches="tight")
plt.close(fig)
print("  → fig02 guardada")

# ── Fig 3: Varianza intra-sujeto del reporte (SD, rango, derivada) ────────
fig, axes = plt.subplots(2, 3, figsize=(16, 10))
# Valencia
boxplot_panel(axes[0,0], val_df, "rep_sd",
              "Valencia – SD intra-sujeto", "DE (escala SAM)")
boxplot_panel(axes[0,1], val_df, "rep_range",
              "Valencia – Rango intra-sujeto", "Rango (escala SAM)")
boxplot_panel(axes[0,2], val_df, "rep_deriv_mean",
              "Valencia – Velocidad media de cambio",
              "Derivada media (SAM/s)")
# Arousal
boxplot_panel(axes[1,0], aro_df, "rep_sd",
              "Arousal – SD intra-sujeto", "DE (escala SAM)")
boxplot_panel(axes[1,1], aro_df, "rep_range",
              "Arousal – Rango intra-sujeto", "Rango (escala SAM)")
boxplot_panel(axes[1,2], aro_df, "rep_deriv_mean",
              "Arousal – Velocidad media de cambio",
              "Derivada media (SAM/s)")
plt.tight_layout(rect=[0, 0.05, 1, 1])
add_legend(fig)
fig.savefig(os.path.join(FIGURES_DIR, "fig03_report_variability.png"),
            dpi=300, bbox_inches="tight")
plt.close(fig)
print("  → fig03 guardada")

# ── Fig 4: ICC temporal por estímulo (scatter ICC vs CV, faceted por dim) ─
fig, axes = plt.subplots(1, 2, figsize=(13, 5.5))
fig.suptitle("ICC Temporal por Estímulo\n"
             "(eje X = CV entre sujetos; color = cuadrante; "
             "línea punteada = referencia ICC=0.70)",
             fontsize=12, fontweight="bold")

for ax, dim_name in zip(axes, ["valence", "arousal"]):
    icc_sub = icc_df[icc_df["dimension"] == dim_name].copy()
    icc_sub["quadrant"] = icc_sub["stimulus"].map(STIM_QUADRANT)
    # CV del estímulo (de nivel 2)
    agg_src = stim_agg_val if dim_name == "valence" else stim_agg_aro
    cv_col  = [c for c in agg_src.columns if "CV entre sujetos" in c]

    for q in QUAD_ORDER:
        sub_q = icc_sub[icc_sub["quadrant"] == q]
        for _, r in sub_q.iterrows():
            cv_val = np.nan
            if cv_col:
                row_agg = agg_src[agg_src["Estímulo"] == r["stimulus"]]
                if len(row_agg):
                    cv_val = row_agg[cv_col[0]].values[0]
            ax.scatter(cv_val, r["icc_sb"],
                       color=QUAD_COLORS[q], s=80, alpha=0.85,
                       edgecolors="white", linewidths=0.5)
            ax.annotate(str(int(r["stimulus"])),
                        (cv_val, r["icc_sb"]),
                        fontsize=8, ha="center", va="bottom",
                        xytext=(0, 4), textcoords="offset points")

    ax.axhline(0.70, color="gray", linestyle="--", linewidth=1)
    ax.axhline(0.0,  color="black", linestyle="-",  linewidth=0.5)
    ax.set_xlabel("CV entre sujetos (variabilidad intrínseca)", fontsize=10)
    ax.set_ylabel("ICC temporal (SB-adj)", fontsize=10)
    ax.set_title(dim_name.capitalize(), fontsize=11, fontweight="bold")
    ax.set_ylim(-0.1, 1.05)

plt.tight_layout(rect=[0, 0.05, 1, 1])
add_legend(fig)
fig.savefig(os.path.join(FIGURES_DIR, "fig04_icc_temporal.png"),
            dpi=300, bbox_inches="tight")
plt.close(fig)
print("  → fig04 guardada")

# ── Fig 5: Espacio circumplejo (scatter Valencia × Arousal) ──────────────
stim_val_mean = (val_df.groupby(["subject_id","stimulus","quadrant"])
                        ["rep_mean"].mean().reset_index()
                        .rename(columns={"rep_mean":"val_score"}))
stim_aro_mean = (aro_df.groupby(["subject_id","stimulus","quadrant"])
                        ["rep_mean"].mean().reset_index()
                        .rename(columns={"rep_mean":"aro_score"}))
scatter_df = pd.merge(stim_val_mean, stim_aro_mean,
                      on=["subject_id","stimulus","quadrant"], how="inner")

if len(scatter_df) > 0:
    fig, ax = plt.subplots(figsize=(7.5, 6.5))
    for q in QUAD_ORDER:
        sub = scatter_df[scatter_df["quadrant"] == q]
        ax.scatter(sub["val_score"], sub["aro_score"],
                   color=QUAD_COLORS[q], alpha=0.45, s=22,
                   label=QUAD_LABELS[q])
    # medias por cuadrante
    for q in QUAD_ORDER:
        sub = scatter_df[scatter_df["quadrant"] == q]
        ax.scatter(sub["val_score"].mean(), sub["aro_score"].mean(),
                   color=QUAD_COLORS[q], s=180, marker="D",
                   edgecolors="black", linewidths=1.2, zorder=5)
    ax.axhline(5, color="gray", linewidth=0.8, linestyle="--")
    ax.axvline(5, color="gray", linewidth=0.8, linestyle="--")
    ax.set_xlim(1, 9)
    ax.set_ylim(1, 9)
    ax.set_xlabel("Autorreporte Valencia", fontsize=11)
    ax.set_ylabel("Autorreporte Arousal", fontsize=11)
    ax.legend(title="Cuadrantes teóricos", title_fontsize=10, fontsize=9, frameon=False)
    plt.tight_layout()
    fig.savefig(os.path.join(FIGURES_DIR, "fig05_circumplex.png"),
                dpi=300, bbox_inches="tight")
    plt.close(fig)
    print("  → fig05 guardada")

# ── Fig 6: Componentes EDA ────────────────────────────────────────────────
fig, axes = plt.subplots(1, 2, figsize=(13, 5.5))
boxplot_panel(axes[0], df, "tonic_mean",  "Tónico – Media (z)",  "z-score")
boxplot_panel(axes[1], df, "phasic_mean", "Fásico – Media (z)",  "z-score")
plt.tight_layout(rect=[0, 0.05, 1, 1])
add_legend(fig)
fig.savefig(os.path.join(FIGURES_DIR, "fig06_eda_components.png"),
            dpi=300, bbox_inches="tight")
plt.close(fig)
print("  → fig06 guardada")

# ── Fig 7: SCR ────────────────────────────────────────────────────────────
fig, axes = plt.subplots(1, 3, figsize=(17, 5.5))
fig.suptitle("Respuestas de Conductancia de la Piel (SCR) por Cuadrante",
             fontsize=13, fontweight="bold")
boxplot_panel(axes[0], df, "scr_amp_mean",     "Amplitud SCR (z)", "z-score")
boxplot_panel(axes[1], df, "scr_rise_mean",    "Tiempo de Subida SCR",     "s")
boxplot_panel(axes[2], df, "scr_rate_per_min", "Frecuencia SCR",   "peaks/min")
plt.tight_layout(rect=[0, 0.05, 1, 1])
add_legend(fig)
fig.savefig(os.path.join(FIGURES_DIR, "fig07_scr.png"),
            dpi=300, bbox_inches="tight")
plt.close(fig)
print("  → fig07 guardada")

# ── Fig 8: SMNA y Potencia Simpática ─────────────────────────────────────
fig, axes = plt.subplots(1, 3, figsize=(17, 5.5))
fig.suptitle("Actividad del Nervio Sudomotor y Potencia Simpática",
             fontsize=13, fontweight="bold")
boxplot_panel(axes[0], df, "smna_auc",       "SMNA – AUC (z·s)",         "z·s")
boxplot_panel(axes[1], df, "smna_amp_mean",  "SMNA – Amplitud media",     "µS")
boxplot_panel(axes[2], df, "sym_band_power", "Potencia Simpática 0.045–0.25 Hz",
              "z²/Hz")
plt.tight_layout(rect=[0, 0.05, 1, 1])
add_legend(fig)
fig.savefig(os.path.join(FIGURES_DIR, "fig08_smna_sym.png"),
            dpi=300, bbox_inches="tight")
plt.close(fig)
print("  → fig08 guardada")

# ── Fig 9: Heatmap EDA por cuadrante ─────────────────────────────────────
key_vars = [
    ("tonic_mean",       "Tónico\nMedia (z)"),
    ("phasic_mean",      "Fásico\nMedia (z)"),
    ("scr_amp_mean",     "Amplitud\nSCR (z)"),
    ("scr_rate_per_min", "Frecuencia\nSCR/min"),
    ("smna_auc",         "SMNA\nAUC"),
    ("sym_band_power",   "Pot.\nSimpática"),
    ("eda_range",        "Rango\nEDA (µS)"),
]
hm_raw = np.array([[df[df["quadrant"]==q][col].mean() for q in QUAD_ORDER]
                    for col, _ in key_vars])
row_m = np.nanmean(hm_raw, axis=1, keepdims=True)
row_s = np.nanstd( hm_raw, axis=1, keepdims=True)
hm_z  = np.where(row_s > 0, (hm_raw - row_m) / row_s, 0.0)

fig, ax = plt.subplots(figsize=(9, 5.5))
im = ax.imshow(hm_z, cmap="RdYlGn", aspect="auto", vmin=-2, vmax=2)
ax.set_xticks(range(len(QUAD_ORDER)))
ax.set_xticklabels([QUAD_LABELS[q].replace(" – ","\n") for q in QUAD_ORDER],
                   fontsize=10)
ax.set_yticks(range(len(key_vars)))
ax.set_yticklabels([lbl for _, lbl in key_vars], fontsize=10)
for i in range(hm_z.shape[0]):
    for j in range(hm_z.shape[1]):
        ax.text(j, i, f"{hm_raw[i,j]:.2f}", ha="center", va="center",
                fontsize=8, color="black")
plt.colorbar(im, ax=ax, label="z-score entre cuadrantes")
ax.set_title("Perfil EDA por Cuadrante\n"
             "(valores crudos en celdas; color = z normalizado entre cuadrantes)",
             fontsize=11, fontweight="bold", pad=10)
plt.tight_layout()
fig.savefig(os.path.join(FIGURES_DIR, "fig09_eda_heatmap.png"),
            dpi=300, bbox_inches="tight")
plt.close(fig)
print("  → fig09 guardada")

# ── Fig 10: Distribución de N y duraciones por sujeto ────────────────────
fig, axes = plt.subplots(1, 2, figsize=(13, 5))
fig.suptitle("Overview del Dataset", fontsize=13, fontweight="bold")

# N registros por sujeto
ax = axes[0]
subj_counts = df.groupby("subject_id").size().sort_values()
ax.barh(subj_counts.index, subj_counts.values,
        color="#5DADE2", alpha=0.85, edgecolor="white")
ax.set_xlabel("N registros", fontsize=10)
ax.set_ylabel("Sujeto", fontsize=10)
ax.set_title("Registros por Sujeto", fontsize=11, fontweight="bold")
ax.axvline(subj_counts.median(), color="gray",
           linestyle="--", linewidth=1, label=f"Mediana={subj_counts.median():.0f}")
ax.legend(fontsize=9, frameon=False)

# Distribución de duraciones
ax = axes[1]
for q in QUAD_ORDER:
    vals = df[df["quadrant"]==q]["duration_s"].dropna().values
    ax.hist(vals, bins=20, alpha=0.55, label=QUAD_LABELS[q],
            color=QUAD_COLORS[q], edgecolor="white")
ax.set_xlabel("Duración del estímulo (s)", fontsize=10)
ax.set_ylabel("Frecuencia", fontsize=10)
ax.set_title("Distribución de Duraciones", fontsize=11, fontweight="bold")
ax.legend(fontsize=9, frameon=False)

plt.tight_layout()
fig.savefig(os.path.join(FIGURES_DIR, "fig10_dataset_overview.png"),
            dpi=300, bbox_inches="tight")
plt.close(fig)
print("  → fig10 guardada")

# ── Fig 11: Medias y Desviaciones por Estímulo ───────────────────────────
fig, axes = plt.subplots(2, 1, figsize=(18, 10), sharex=True)
fig.suptitle("Promedio y Desviación Estándar del Reporte por Estímulo",
             fontsize=14, fontweight="bold")

# Datos de Valencia
val_plot_df = stim_agg_val[["Estímulo", "Cuadrante", "Valencia (escala 1–9) – mean", "Valencia (escala 1–9) – sd"]].copy()
val_plot_df.columns = ["stimulus", "quadrant", "mean", "sd"]
val_plot_df = val_plot_df.sort_values("stimulus")

# Datos de Arousal
aro_plot_df = stim_agg_aro[["Estímulo", "Cuadrante", "Arousal (escala 1–9) – mean", "Arousal (escala 1–9) – sd"]].copy()
aro_plot_df.columns = ["stimulus", "quadrant", "mean", "sd"]
aro_plot_df = aro_plot_df.sort_values("stimulus")

# Panel Valencia
ax = axes[0]
bar_colors = val_plot_df["quadrant"].map(QUAD_COLORS)
ax.bar(val_plot_df["stimulus"].astype(str), val_plot_df["mean"],
       yerr=val_plot_df["sd"], capsize=4, color=bar_colors, alpha=0.8)
ax.axhline(5, color="gray", linestyle="--", linewidth=1)
ax.set_ylabel("Valencia (escala SAM 1-9)", fontsize=10)
ax.set_title("Valencia", fontsize=12, fontweight="bold")
ax.grid(axis='y', linestyle='--', alpha=0.6)

# Panel Arousal
ax = axes[1]
bar_colors = aro_plot_df["quadrant"].map(QUAD_COLORS)
ax.bar(aro_plot_df["stimulus"].astype(str), aro_plot_df["mean"],
       yerr=aro_plot_df["sd"], capsize=4, color=bar_colors, alpha=0.8)
ax.axhline(5, color="gray", linestyle="--", linewidth=1)
ax.set_ylabel("Arousal (escala SAM 1-9)", fontsize=10)
ax.set_title("Arousal", fontsize=12, fontweight="bold")
ax.grid(axis='y', linestyle='--', alpha=0.6)

ax.set_xlabel("ID del Estímulo", fontsize=10)
plt.xticks(rotation=90, fontsize=8)
add_legend(fig)
plt.tight_layout(rect=[0, 0.05, 1, 0.96])
fig.savefig(os.path.join(FIGURES_DIR, "fig11_stimulus_means.png"),
            dpi=300, bbox_inches="tight")
plt.close(fig)
print("  → fig11 guardada")


# Helper para plots de trayectoria 2D
def plot_2d_trajectory(ax, stim_id, title, raw_profiles_dict):
    """Dibuja la trayectoria Valencia vs Arousal para un estímulo y dos sujetos."""
    ax.set_title(title, fontsize=11, fontweight="bold")
    subjects_with_both = []
    
    val_key = (stim_id, "valence")
    aro_key = (stim_id, "arousal")
    
    if val_key not in raw_profiles_dict or aro_key not in raw_profiles_dict:
        ax.text(0.5, 0.5, "Datos insuficientes", ha="center", va="center")
        return

    val_subjects = set(raw_profiles_dict[val_key].keys())
    aro_subjects = set(raw_profiles_dict[aro_key].keys())
    common_subjects = sorted(list(val_subjects.intersection(aro_subjects)))
    
    if len(common_subjects) < 2:
        ax.text(0.5, 0.5, f"Menos de 2 sujetos\ncomunes ({len(common_subjects)})", ha="center", va="center")
        return

    subjects_to_plot = common_subjects[:2]
    colors = ['#3498DB', '#E74C3C']
    
    for i, sid in enumerate(subjects_to_plot):
        val_trace = raw_profiles_dict[val_key][sid]
        aro_trace = raw_profiles_dict[aro_key][sid]
        
        # Truncar a la longitud más corta
        min_len = min(len(val_trace), len(aro_trace))
        val_trace = val_trace[:min_len]
        aro_trace = aro_trace[:min_len]
        
        # El eje X es valencia, el eje Y es arousal (ejes del joystick)
        ax.plot(val_trace, aro_trace, color=colors[i], alpha=0.7, label=f'Sujeto {sid}')
        ax.scatter(val_trace[0], aro_trace[0], color=colors[i], marker='o', s=50, ec='k', zorder=5, label=f'Inicio S{sid}')
        ax.scatter(val_trace[-1], aro_trace[-1], color=colors[i], marker='X', s=50, ec='k', zorder=5, label=f'Fin S{sid}')

    ax.axhline(5, color="k", linestyle="--", linewidth=0.7, alpha=0.5)
    ax.axvline(5, color="k", linestyle="--", linewidth=0.7, alpha=0.5)
    ax.set_xlim(1, 9)
    ax.set_ylim(1, 9)
    ax.set_xlabel("Valencia (escala SAM 1-9)", fontsize=10)
    ax.set_ylabel("Arousal (escala SAM 1-9)", fontsize=10)
    ax.legend(fontsize=7)
    ax.grid(True, linestyle=':', alpha=0.6)
    ax.set_aspect('equal', adjustable='box')


# ── Fig 12: Trayectorias 2D (Valencia vs Arousal) para varios sujetos ────
STIMULUS_ID_2D = 9               # Hardcodea aquí el ID del estímulo
SUBJECTS_2D = ["19", "20", "21"] # Hardcodea aquí los IDs de los sujetos

fig, ax = plt.subplots(figsize=(7, 7))
fig.suptitle(f"Trayectorias 2D de Reporte Continuo - Estímulo {STIMULUS_ID_2D}\n"
             "(Señales crudas, sin suavizar)", fontsize=13, fontweight="bold")

val_key = (STIMULUS_ID_2D, "valence")
aro_key = (STIMULUS_ID_2D, "arousal")

# Extraer diccionarios de perfiles crudos
val_profiles = raw_joy_profiles.get(val_key, {})
aro_profiles = raw_joy_profiles.get(aro_key, {})

colors_2d = ['#3498DB', '#E74C3C', '#2ECC71', '#9B59B6', '#F39C12']
plotted_any = False

for i, sid in enumerate(SUBJECTS_2D):
    # Verificamos que el sujeto tenga ambos reportes para este estímulo
    if sid in val_profiles and sid in aro_profiles:
        v_trace = val_profiles[sid]
        a_trace = aro_profiles[sid]
        
        # Truncar a la longitud más corta para emparejar ambas dimensiones
        min_len = min(len(v_trace), len(a_trace))
        v_trace = v_trace[:min_len]
        a_trace = a_trace[:min_len]
        
        c = colors_2d[i % len(colors_2d)]
        
        # Dibujar trayectoria
        ax.plot(v_trace, a_trace, color=c, alpha=0.75, linewidth=1.5, label=f'Sujeto {sid}')
        
        # Marcar inicio (círculo) y fin (cruz)
        ax.scatter(v_trace[0], a_trace[0], color=c, marker='o', s=50, ec='k', zorder=5)
        ax.scatter(v_trace[-1], a_trace[-1], color=c, marker='X', s=60, ec='k', zorder=5)
        
        plotted_any = True
    else:
        print(f"  ⚠ Fig 12: Faltan datos de Valencia o Arousal para el sujeto {sid} en estímulo {STIMULUS_ID_2D}")

if not plotted_any:
    ax.text(0, 0, "No hay datos suficientes\npara los sujetos indicados", ha="center", va="center")

# Líneas centrales y límites fijos en el rango del joystick
ax.axhline(0, color="k", linestyle="--", linewidth=0.7, alpha=0.5)
ax.axvline(0, color="k", linestyle="--", linewidth=0.7, alpha=0.5)
ax.set_xlim(-1.1, 1.1)
ax.set_ylim(-1.1, 1.1)
ax.set_xlabel("Joystick Valencia (-1 a 1)", fontsize=10)
ax.set_ylabel("Joystick Arousal (-1 a 1)", fontsize=10)

# Añadir leyenda que incluya la explicación de los marcadores
import matplotlib.lines as mlines
handles, labels = ax.get_legend_handles_labels()
if plotted_any:
    handles.append(mlines.Line2D([], [], color='gray', marker='o', linestyle='None', label='Inicio'))
    handles.append(mlines.Line2D([], [], color='gray', marker='X', linestyle='None', label='Fin'))
ax.legend(handles=handles, fontsize=9, loc="upper right", framealpha=0.9)

ax.grid(True, linestyle=':', alpha=0.6)
ax.set_aspect('equal', adjustable='box')

plt.tight_layout()
fig.savefig(os.path.join(FIGURES_DIR, "fig12_2d_trajectories.png"), dpi=300, bbox_inches="tight")
plt.close(fig)
print("  → fig12 guardada")

# ── Fig 13: Perfil temporal promedio para un estímulo específico ─────────
STIMULUS_IDS_FOR_PLOT = [n for n in range(15)]  # Hardcodea aquí el ID del estímulo que quieres visualizar

def plot_mean_stimulus_profile(ax, stim_id, dimension, data_profiles, color, title):
    """
    Calcula y plotea el perfil temporal promedio y su DE para un estímulo.
    Los perfiles se remuestrean a una grilla común para poder promediarlos.
    """
    key = (stim_id, dimension)
    if key not in data_profiles or not data_profiles[key]:
        ax.text(0.5, 0.5, "No hay datos disponibles", ha="center", va="center", fontsize=12)
        ax.set_title(title, fontsize=12, fontweight="bold")
        return

    profiles_dict = data_profiles[key]
    
    # Remuestreo de todos los perfiles a una grilla temporal común
    grid = np.linspace(0, 100, ICC_GRID_POINTS) # Eje X como % de duración
    resampled_profiles = []
    for sid, profile in profiles_dict.items():
        if len(profile) < 4: continue
        x_original = np.linspace(0, 100, len(profile))
        try:
            f = interp1d(x_original, profile, kind='linear', fill_value="extrapolate")
            resampled_profiles.append(f(grid))
        except Exception:
            continue
    
    if len(resampled_profiles) < 2:
        ax.text(0.5, 0.5, "Datos insuficientes para promediar", ha="center", va="center", fontsize=12)
        ax.set_title(title, fontsize=12, fontweight="bold")
        return

    # Cálculo de la media y la desviación estándar en cada punto temporal
    matrix = np.vstack(resampled_profiles)
    mean_profile = np.nanmean(matrix, axis=0)

    # 1. Ploteo de trayectorias individuales en el fondo
    # Usamos un mapa de colores (colormap) para darle un color distinto a cada sujeto
    cmap = plt.cm.get_cmap('tab20', len(resampled_profiles))
    
    for i, profile_array in enumerate(resampled_profiles):
        # alpha=0.35 da la transparencia para que no tapen a la media
        ax.plot(grid, profile_array, color=cmap(i), alpha=0.35, linewidth=1.2)

    # 2. Ploteo de la trayectoria promedio por encima (sólida y en negro)
    ax.plot(grid, mean_profile, color='black', linewidth=3.0, alpha=1.0, zorder=10, 
            label=f'Media (N={len(resampled_profiles)})')
    
    ax.axhline(5, color='gray', linestyle='--', linewidth=1, alpha=0.8)
    ax.set_ylim(1, 9)
    ax.set_ylabel("Valor del reporte", fontsize=10)
    ax.set_title(title, fontsize=12, fontweight="bold")
    ax.legend()
    ax.grid(True, linestyle=':', alpha=0.7)

for n in STIMULUS_IDS_FOR_PLOT:
    if (n, "valence") not in joy_profiles or (n, "arousal") not in joy_profiles:
        print(f"  ⚠ Fig 13: No hay datos de perfiles para el estímulo {n} en una o ambas dimensiones")
        continue

    fig, axes = plt.subplots(2, 1, figsize=(12, 8), sharex=True)

    # Panel superior: Valencia
    plot_mean_stimulus_profile(axes[0], n, 'valence', joy_profiles,
                            color='#3498DB', title='Valencia')

    # Panel inferior: Arousal
    plot_mean_stimulus_profile(axes[1], n, 'arousal', joy_profiles,
                            color='#E74C3C', title='Arousal')

    axes[1].set_xlabel("% de Duración del Estímulo", fontsize=11)

    plt.tight_layout(rect=[0, 0, 1, 0.95])
    fig.savefig(os.path.join(FIGURES_DIR, f"fig13_mean_stimulus_profile_{n}.png"),
                dpi=300, bbox_inches="tight")
    plt.close(fig)
    print(f"  → fig13 estímulo {n} guardada")

# ── Fig 14: Distribución de componentes EDA — Sesión A vs B ──────────────
print("  → Generando fig14...")

eda_components = [
    ("tonic_mean",       "Tónico Media (z)"),
    ("phasic_mean",      "Fásico Media (z)"),
    ("scr_amp_mean",     "Amplitud SCR (z)"),
    ("scr_rate_per_min", "Frecuencia SCR (peaks/min)"),
    ("smna_auc",         "SMNA AUC (z·s)"),
    ("sym_band_power",   "Potencia Simpática (z²/Hz)"),
]

# Extraer letra de sesión (A o B) desde el campo 'session'
# Ajustá esta línea si tu nomenclatura es distinta (ej: "VR1" → tomar último char)
print(df["session"].unique())

df["session_ab"] = df["session"].str[-1].str.upper()
sessions = sorted(df["session_ab"].unique())  # ['A', 'B'] o lo que haya

if len(sessions) < 2:
    print("    ⚠ Solo hay una sesión en los datos, fig14 omitida.")
else:
    n_cols = 2
    n_rows = len(eda_components)
    fig, axes = plt.subplots(n_rows, n_cols, figsize=(13, n_rows * 3.5))
    fig.suptitle("Distribución de Componentes EDA: Sesión A vs Sesión B\n"
                 "(un punto = registro sujeto × estímulo)",
                 fontsize=14, fontweight="bold")

    ses_colors = {"A": "#5DADE2", "B": "#E59866"}

    for row_idx, (col, label) in enumerate(eda_components):
        if col not in df.columns:
            continue

        # Panel izquierdo: boxplot A vs B
        ax_box = axes[row_idx, 0]
        groups = [df[df["session_ab"] == s][col].dropna().values for s in sessions]
        bp = ax_box.boxplot(groups, patch_artist=True,
                            medianprops=dict(color="black", linewidth=2),
                            whiskerprops=dict(linewidth=1.2),
                            capprops=dict(linewidth=1.2),
                            flierprops=dict(marker="o", markersize=3, alpha=0.4))
        for patch, s in zip(bp["boxes"], sessions):
            patch.set_facecolor(ses_colors.get(s, "#AAAAAA"))
            patch.set_alpha(0.80)
        for i, (g, s) in enumerate(zip(groups, sessions), 1):
            jit = np.random.normal(i, 0.07, size=len(g))
            ax_box.scatter(jit, g, color=ses_colors.get(s, "#AAAAAA"),
                           alpha=0.35, s=12, zorder=3)
        ax_box.set_xticks(range(1, len(sessions) + 1))
        ax_box.set_xticklabels([f"Sesión {s}" for s in sessions], fontsize=10)
        ax_box.set_ylabel(label, fontsize=9)
        ax_box.set_title(f"{label} — Boxplot", fontsize=10, fontweight="bold")

        # Panel derecho: histograma superpuesto A vs B
        ax_hist = axes[row_idx, 1]
        for s in sessions:
            vals = df[df["session_ab"] == s][col].dropna().values
            ax_hist.hist(vals, bins=30, alpha=0.55,
                         color=ses_colors.get(s, "#AAAAAA"),
                         label=f"Sesión {s} (N={len(vals)})",
                         edgecolor="white", density=True)
        ax_hist.set_xlabel(label, fontsize=9)
        ax_hist.set_ylabel("Densidad", fontsize=9)
        ax_hist.set_title(f"{label} — Histograma", fontsize=10, fontweight="bold")
        ax_hist.legend(fontsize=8, frameon=False)

    plt.tight_layout(rect=[0, 0, 1, 0.96])
    fig.savefig(os.path.join(FIGURES_DIR, "fig14_eda_session_ab.png"),
                dpi=300, bbox_inches="tight")
    plt.close(fig)
    print("  → fig14 guardada")

# ── Fig 15: Perfil temporal promedio de EDA Clean (Estímulo 13, Sesión A) ──────────
print("  → Generando fig15...")
for STIM_EDA in range(1, 15):
    STIM_EDA_PLOT = STIM_EDA
    TARGET_SESSION = 'A'

    key_eda_target = (STIM_EDA_PLOT, TARGET_SESSION)

    if key_eda_target not in eda_profiles:
        print(f"    ⚠ fig15: No se encontraron datos para el Estímulo {STIM_EDA_PLOT} en la Sesión {TARGET_SESSION}.")
    else:
        target_eda_data = eda_profiles[key_eda_target]
        
        # 1. Verificar sujetos faltantes e imprimir advertencia
        missing_subjects = [sid for sid in SUBJECT_IDS if sid not in target_eda_data]
        if missing_subjects:
            print(f"    ⚠ fig15: Los siguientes sujetos NO tienen el estímulo {STIM_EDA_PLOT} en la Sesión {TARGET_SESSION} y serán omitidos:")
            print(f"             {missing_subjects}")

        # 2. Remuestreo a grilla común (0 a 100% de la duración)
        grid_eda = np.linspace(0, 100, ICC_GRID_POINTS)
        resampled_eda = []
        
        for sid, profile in target_eda_data.items():
            if len(profile) < 4:
                continue
            x_original = np.linspace(0, 100, len(profile))
            try:
                f = interp1d(x_original, profile, kind='linear', fill_value="extrapolate")
                resampled_eda.append(f(grid_eda))
            except Exception:
                continue
                
        # 3. Calcular medias y graficar
        if len(resampled_eda) < 2:
            print(f"    ⚠ fig15: Datos insuficientes (<2 sujetos) para promediar.")
        else:
            matrix_eda = np.vstack(resampled_eda)
            mean_eda = np.nanmean(matrix_eda, axis=0)
            std_eda = np.nanstd(matrix_eda, axis=0, ddof=1)
            
            fig, ax = plt.subplots(figsize=(10, 5.5))
            ax.plot(grid_eda, mean_eda, color='#27AE60', linewidth=2.5, 
                    label=f'Media (N={len(resampled_eda)})')
            ax.fill_between(grid_eda, mean_eda - std_eda, mean_eda + std_eda,
                            color='#27AE60', alpha=0.2, label='±1 Desv. Estándar')
            
            ax.axhline(0, color='gray', linestyle='--', linewidth=1, alpha=0.8)
            ax.set_xlabel("% de Duración del Estímulo", fontsize=11)
            ax.set_ylabel("EDA Clean (z-score)", fontsize=11)
            ax.set_title(f"Perfil Temporal Promedio de EDA - Estímulo {STIM_EDA_PLOT} (Sesión {TARGET_SESSION})\n"
                        "(Línea = media entre sujetos; Sombra = ±1 DE)",
                        fontsize=13, fontweight="bold")
            ax.legend(loc="upper left")
            ax.grid(True, linestyle=':', alpha=0.7)
            
            plt.tight_layout()
            fig.savefig(os.path.join(FIGURES_DIR, f"fig15_mean_eda_profile_stim{STIM_EDA_PLOT}.png"),
                        dpi=300, bbox_inches="tight")
            plt.close(fig)
            print("  → fig15 guardada")

# ─────────────────────────────────────────────────────────────────────────────
# Fig 16: Componentes principales EDA por cuadrante (2 arriba, 1 abajo)
# ─────────────────────────────────────────────────────────────────────────────

fig16_vars = [
    ("tonic_mean",  "Tónico",        "z"),
    ("phasic_mean", "Fásico",         "z"),
    ("smna_auc",    "AUC de SMNA",  "z"),
]

# --- CAMBIO 1: Definir el layout con subplot_mosaic ---
# 'A' y 'B' estarán en la primera fila.
# 'C' estará en la segunda fila, ocupando el mismo espacio horizontal.
layout = [
    ["A", "B"],
    ["C", "C"],
]
fig16, axes16 = plt.subplot_mosaic(layout, figsize=(12, 10))

# --- CAMBIO 2: Iterar sobre los gráficos y ejes específicos ---
# Gráfico A
boxplot_panel(axes16['A'], df, fig16_vars[0][0], fig16_vars[0][1], fig16_vars[0][2])
# Gráfico B
boxplot_panel(axes16['B'], df, fig16_vars[1][0], fig16_vars[1][1], fig16_vars[1][2])
# Gráfico C
boxplot_panel(axes16['C'], df, fig16_vars[2][0], fig16_vars[2][1], fig16_vars[2][2])


plt.tight_layout(rect=[0, 0.08, 1, 1])
fig16.savefig(
    os.path.join(FIGURES_DIR, "fig16_eda_main_components.png"),
    dpi=300, bbox_inches="tight",
)
plt.close(fig16)
print("  → fig16 guardada")

# ─────────────────────────────────────────────────────────────────────────────
# Fig 17: Features extraídas de SCR y SMNA por cuadrante (2 arriba, 1 abajo)
# ─────────────────────────────────────────────────────────────────────────────

fig17_vars = [
    ("scr_rate_per_min", "Picos SCR",        "Peaks/min"),
    ("smna_count",       "Picos SMNA",    "N picos"),
    ("scr_amp_mean",     "Amplitud SCR",  "z"),
]

# --- CAMBIO 1: Definir el layout con subplot_mosaic ---
layout = [
    ["A", "B"],
    ["C", "C"],
]
fig17, axes17 = plt.subplot_mosaic(layout, figsize=(15, 10))

# --- CAMBIO 2: Iterar sobre los gráficos y ejes específicos ---
# Gráfico A
boxplot_panel(axes17['A'], df, fig17_vars[0][0], fig17_vars[0][1], fig17_vars[0][2])
# Gráfico B
boxplot_panel(axes17['B'], df, fig17_vars[1][0], fig17_vars[1][1], fig17_vars[1][2])
# Gráfico C
boxplot_panel(axes17['C'], df, fig17_vars[2][0], fig17_vars[2][1], fig17_vars[2][2])

plt.tight_layout(rect=[0, 0.08, 1, 1])
fig17.savefig(
    os.path.join(FIGURES_DIR, "fig17_scr_smna_features.png"),
    dpi=300, bbox_inches="tight",
)
plt.close(fig17)
print("  → fig17 guardada")

# ── Fig 18: Espacio circumplejo con medias por estímulo ──────────────────
# Nube de fondo : un punto por sujeto × estímulo (igual que Fig 5).
# Marcadores    : uno por estímulo (N=14).
#   - Forma  → cuadrante teórico (Russell 1980)
#   - Color  → ID del estímulo (paleta de 14 colores distintos)
#   - Label  → ID del estímulo

if len(scatter_df) > 0:

    # Paleta de 14 colores cualitativos bien separados
    cmap_14 = plt.cm.get_cmap("tab20", 14)
    stim_ids_sorted = sorted(scatter_df["stimulus"].unique())
    stim_color = {sid: cmap_14(i) for i, sid in enumerate(stim_ids_sorted)}

    # Marcador por cuadrante teórico
    quad_marker = {
        "HVHA": "^",   # triángulo arriba
        "HVLA": "s",   # cuadrado
        "LVHA": "D",   # diamante
        "LVLA": "o",   # círculo
    }

    fig, ax = plt.subplots(figsize=(8.5, 7.5))

    # ── Nube de fondo: sujeto × estímulo, coloreada por cuadrante teórico ──
    for q in QUAD_ORDER:
        sub = scatter_df[scatter_df["quadrant"] == q]
        ax.scatter(sub["val_score"], sub["aro_score"],
                   color=QUAD_COLORS[q], alpha=0.25, s=18, zorder=1)

    # ── Media por estímulo ───────────────────────────────────────────────
    # Usamos df directamente para incluir todos los sujetos que vieron cada
    # estímulo, sin requerir que tengan ambas dimensiones en scatter_df.
    val_per_stim = (df[df["dimension"] == "valence"]
                    .groupby("stimulus")["rep_mean"].mean())
    aro_per_stim = (df[df["dimension"] == "arousal"]
                    .groupby("stimulus")["rep_mean"].mean())
    stim_means_18 = (pd.DataFrame({"val": val_per_stim, "aro": aro_per_stim})
                     .dropna()
                     .reset_index())
    stim_means_18["quadrant"] = stim_means_18["stimulus"].map(STIM_QUADRANT)

    for _, r in stim_means_18.iterrows():
        sid  = int(r["stimulus"])
        q    = r["quadrant"]
        mkr  = quad_marker.get(q, "o")
        col  = stim_color[sid]

        ax.scatter(r["val"], r["aro"],
                   color=col, marker=mkr,
                   s=220, edgecolors="black", linewidths=1.1,
                   zorder=5)
        ax.annotate(str(sid),
                    (r["val"], r["aro"]),
                    fontsize=8, fontweight="bold",
                    ha="center", va="bottom",
                    xytext=(0, 7), textcoords="offset points",
                    color="black")

    # ── Ejes y decoraciones ──────────────────────────────────────────────
    ax.axhline(5, color="gray", linewidth=0.8, linestyle="--", alpha=0.7)
    ax.axvline(5, color="gray", linewidth=0.8, linestyle="--", alpha=0.7)
    ax.set_xlim(1, 9)
    ax.set_ylim(1, 9)
    ax.set_xlabel("Autorreporte Valencia (escala SAM 1–9)", fontsize=11)
    ax.set_ylabel("Autorreporte Arousal (escala SAM 1–9)",  fontsize=11)
    ax.set_title(
        "Espacio Circumplejo: Media por Estímulo\n"
        "(nube = sujeto × estímulo; marcador = media del estímulo; "
        "forma = cuadrante teórico; color = ID estímulo)",
        fontsize=11, fontweight="bold",
    )

    # Leyenda 1: cuadrantes teóricos (forma)
    shape_handles = [
        plt.scatter([], [], marker=quad_marker[q], s=100,
                    facecolor="gray", edgecolors="black", linewidths=0.8,
                    label=QUAD_LABELS[q])
        for q in QUAD_ORDER
    ]
    leg1 = ax.legend(handles=shape_handles, title="Cuadrante teórico\n(forma)",
                     fontsize=8, title_fontsize=8,
                     loc="upper left", frameon=True, framealpha=0.85)
    ax.add_artist(leg1)

    # Leyenda 2: estímulos (color)
    color_handles = [
        plt.scatter([], [], marker="o", s=80,
                    color=stim_color[sid], edgecolors="black", linewidths=0.6,
                    label=f"Estímulo {sid}  ({STIM_QUADRANT.get(sid,'?')})")
        for sid in stim_ids_sorted
    ]
    ax.legend(handles=color_handles, title="Estímulo (color)",
              fontsize=7.5, title_fontsize=8,
              loc="lower right", frameon=True, framealpha=0.85,
              ncol=2)

    plt.tight_layout()
    fig.savefig(os.path.join(FIGURES_DIR, "fig18_circumplex_by_stimulus.png"),
                dpi=300, bbox_inches="tight")
    plt.close(fig)
    print("  → fig18 guardada")

# ── Fig 19: Autorreporte promedio (Valencia y Arousal) ───────────────────
fig, axes = plt.subplots(1, 2, figsize=(13, 5.5))

# Panel izquierdo: Valencia
boxplot_panel(axes[0], val_df, "rep_mean",
              "Valencia", "Reporte promedio")
axes[0].axhline(5, color="gray", linewidth=0.8, linestyle="--", alpha=0.6)

# Panel derecho: Arousal
boxplot_panel(axes[1], aro_df, "rep_mean",
              "Arousal", "Reporte promedio")
axes[1].axhline(5, color="gray", linewidth=0.8, linestyle="--", alpha=0.6)

plt.tight_layout(rect=[0, 0.05, 1, 1])
fig.savefig(os.path.join(FIGURES_DIR, "fig19_val_aro_means.png"),
            dpi=300, bbox_inches="tight")
plt.close(fig)
print("  → fig19 guardada")

# ─────────────────────────────────────────────────────────────────────────────
# RESUMEN FINAL
# ─────────────────────────────────────────────────────────────────────────────
print(f"\n{'='*60}")
print("  OUTPUTS GENERADOS")
print(f"{'='*60}")
print(f"\nDirectorio: {OUTPUT_DIR}\n")

all_files = []
for root, dirs, files in os.walk(OUTPUT_DIR):
    for f in sorted(files):
        all_files.append(os.path.join(root, f))

for fp in sorted(all_files):
    size_kb = os.path.getsize(fp) / 1024
    print(f"  {fp:<65} ({size_kb:.1f} KB)")

print(f"\n{'='*60}")
print(f"  Total registros nivel 1: {len(df)}")
print(f"  Sujetos incluidos:       {df['subject_id'].nunique()}")
print(f"  Estímulos únicos:        {df['stimulus'].nunique()}")
print(f"  Figuras generadas:       17+n_estimulos*2 (figs 01–15, 300 dpi)")
print(f"{'='*60}\n")